from datetime import datetime, timedelta
import hashlib
import io
import os
import re
import tempfile
from fpdf import FPDF
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import qrcode
import streamlit as st

import db

def pdf_sanitize(text):
    text = str(text)
    replacements = {
        'ą':'a', 'ć':'c', 'ę':'e', 'ł':'l', 'ń':'n', 'ó':'o', 'ś':'s', 'ź':'z', 'ż':'z',
        'Ą':'A', 'Ć':'C', 'Ę':'E', 'Ł':'L', 'Ń':'N', 'Ó':'O', 'Ś':'S', 'Ź':'Z', 'Ż':'Z',
        '€':'EUR', '–':'-', '—':'-', '”':'"', '„':'"', '’':"'", '“':'"', '\xa0':' '
    }
    for pl, eng in replacements.items():
        text = text.replace(pl, eng)
    return text.encode('latin-1', 'ignore').decode('latin-1')

class PRO_TransportOrder(FPDF):
    def __init__(self, watermark_text="SQM", opiekun="PD"):
        super().__init__()
        self.watermark_text = pdf_sanitize(watermark_text)
        self.opiekun = opiekun
        self.primary_color = (25, 118, 210) 
        self.dark_text = (40, 40, 40)
        self.light_text = (100, 100, 100)

    def add_watermark(self):
        self.set_font("Arial", 'B', 45)
        self.set_text_color(245, 245, 245) 
        for j in range(80, 297, 45):
            przesuniecie = 35 if (j // 45) % 2 == 0 else 0
            for i in range(-20, 210, 70):
                self.text(i + przesuniecie, j, self.watermark_text)
        self.set_text_color(0, 0, 0)

    def header(self):
        try:
            if os.path.exists("logosqm.png"):
                self.image("logosqm.png", 10, 6, 57.5)
            elif os.path.exists("logosqm.jpg"):
                self.image("logosqm.jpg", 10, 6, 57.5)
        except:
            pass
        
        self.set_font("Arial", 'B', 18)
        self.set_text_color(*self.dark_text)
        self.set_xy(65, 10)
        self.cell(105, 8, pdf_sanitize("TRANSPORT ORDER"), ln=True, align='R')
        
        self.set_font("Arial", 'B', 11)
        self.set_text_color(*self.light_text)
        self.set_xy(65, 18)
        self.cell(105, 5, pdf_sanitize("ZLECENIE TRANSPORTOWE"), ln=True, align='R')
        
        self.set_font("Arial", '', 8)
        self.set_text_color(*self.light_text)
        self.set_xy(65, 23)
        self.cell(105, 5, pdf_sanitize("SQM Prosta Spółka Akcyjna | Logistics Department"), ln=True, align='R')

    def footer(self):
        self.set_y(-30)
        self.set_font("Arial", 'I', 10)
        self.set_text_color(*self.primary_color)
        self.cell(0, 5, pdf_sanitize("Thank you for your cooperation! / Dziękujemy za współpracę!"), ln=True, align='C')
        
        self.set_font("Arial", '', 8)
        self.set_text_color(*self.light_text)
        
        if self.opiekun == "PD":
            email = "piotr.dukiel@sqm.eu"
            telefon = "+48 577 63 63 67"
        else:
            email = "piotr.kaczmarek@sqm.eu"
            telefon = "+48 570 33 02 90"
            
        self.cell(0, 5, pdf_sanitize(f"www.sqm.pl   |   {email}   |   {telefon}"), ln=True, align='C')

def generate_pro_pdf(dane):
    pdf = PRO_TransportOrder(opiekun=dane.get('opiekun', 'PD'))
    pdf.alias_nb_pages()
    
    pdf.set_auto_page_break(auto=True, margin=33) 
    
    pdf.add_page()
    pdf.add_watermark()

    token_base = f"{dane['nr']}-{dane['przewoznik_nazwa']}-{dane['stawka']}"
    secure_hash = hashlib.md5(token_base.encode()).hexdigest()[:12].upper()
    qr_content = f"SQM-VERIFY: {dane['nr']}\nVALID-HASH: {secure_hash}\nSYSTEM: SQM HUB"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(qr_content)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
        img_qr.save(tmp, format="PNG")
        qr_path = tmp.name
        
    pdf.image(qr_path, 175, 8, 22)
    if os.path.exists(qr_path): os.remove(qr_path)

    pdf.set_xy(10, 32)
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(25, 118, 210) 
    pdf.set_text_color(255, 255, 255)
    pdf.cell(25, 8, pdf_sanitize(" REF "), border=0, fill=True, align='C')
    pdf.set_fill_color(245, 245, 245)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(60, 8, pdf_sanitize(f" {dane['nr']}"), border=0, fill=True)
    pdf.cell(5, 8, "", border=0) 
    pdf.set_fill_color(25, 118, 210)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(25, 8, pdf_sanitize(" DATE "), border=0, fill=True, align='C')
    pdf.set_fill_color(245, 245, 245)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(60, 8, pdf_sanitize(f" {datetime.now().strftime('%d.%m.%Y')}"), border=0, fill=True)
    
    pdf.ln(6)

    def draw_section_header(num, title):
        pdf.set_fill_color(25, 118, 210)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 11) 
        pdf.cell(8, 8, pdf_sanitize(str(num).zfill(2)), fill=True, align='C')
        pdf.set_text_color(40, 40, 40)
        pdf.cell(3, 8, "", border=0)
        pdf.cell(0, 8, pdf_sanitize(title), ln=True)
        pdf.ln(1)

    def draw_row(label, val, border_b=True):
        x_start = pdf.get_x()
        y_start = pdf.get_y()
        pdf.set_font("Arial", 'B', 7.5) 
        pdf.set_text_color(100, 100, 100)
        pdf.cell(60, 5, pdf_sanitize(label), border=0)
        pdf.set_font("Arial", 'B', 9) 
        pdf.set_text_color(40, 40, 40)
        pdf.set_xy(x_start + 60, y_start + 0.5)
        pdf.multi_cell(130, 4.5, pdf_sanitize(val), border=0)
        y_end = pdf.get_y() + 0.5 
        if border_b:
            pdf.set_draw_color(230, 230, 230)
            pdf.line(10, y_end, 200, y_end)
        pdf.set_xy(10, y_end + 1.5)

    draw_section_header(1, "PARTIES & ASSETS / STRONY I POJAZD")
    draw_row("CONTRACTOR / PRZEWOŹNIK:", dane['przewoznik_detale'])
    draw_row("VEHICLE & DRIVER / AUTO I KIEROWCA:", dane['auto'] if dane['auto'] else "TBA / Do podania")
    draw_row("VALUATION MODEL / TRYB WYCENY:", dane['typ_zlecenia'], border_b=False)
    pdf.ln(2)

    draw_section_header(2, "LOGISTICS TIMELINE / HARMONOGRAM LOGISTYCZNY")
    draw_row("LOADING PLACE / MIEJSCE ZAŁADUNKU:", dane['zaladunek'])
    draw_row("LOADING DATE / DATA ZAŁADUNKU:", dane['data_zal'])
    draw_row("UNLOADING DATE / DATA ROZŁADUNKU:", dane['data_roz'])
    draw_row("UNLOADING PLACE / MIEJSCE ROZŁADUNKU:", dane['rozladunek'])
    
    if dane['typ_zlecenia'] == "Pełny event":
        emp1 = str(dane.get('data_emp_in_1', ''))
        emp2 = str(dane.get('data_emp_in_2', ''))
        emp_str = emp1 if (emp1 and emp1 != 'None') else ''
        if emp2 and emp2 != 'None' and emp2.strip():
            emp_str += f" & {emp2}" if emp_str else emp2
            
        draw_row("EMPTIES IN / ODBIÓR PUSTYCH:", emp_str if emp_str else "---")
        
        dost_pust = str(dane.get('data_dostawa_pustych', ''))
        if dost_pust and dost_pust != 'None' and dost_pust.strip():
            draw_row("EMPTIES DELIVERY / DOSTAWA PUSTYCH:", dost_pust)
            
        odb_peln = str(dane.get('data_odbior_pelnych', '---'))
        if not odb_peln or odb_peln == 'None' or not odb_peln.strip():
            odb_peln = '---'
            
        draw_row("RETURN LOAD / ODBIÓR PEŁNYCH:", odb_peln, border_b=False)
    else:
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)

    draw_section_header(3, "FINANCIALS & CARGO / FINANSE I ŁADUNEK")
    sy = pdf.get_y()
    
    pdf.set_xy(120, sy); pdf.set_fill_color(25, 118, 210); pdf.rect(120, sy, 22, 22, 'F') 
    pdf.set_xy(142, sy); pdf.set_fill_color(25, 118, 210); pdf.rect(142, sy, 58, 22, 'F')
    pdf.set_xy(142, sy + 3); pdf.set_font("Arial", 'B', 8); pdf.set_text_color(255, 255, 255)
    pdf.cell(55, 4, pdf_sanitize("TOTAL NET RATE / KWOTA NETTO"), ln=True)
    pdf.set_xy(142, sy + 9); pdf.set_font("Arial", 'B', 18)
    pdf.cell(55, 9, pdf_sanitize(f"{dane['stawka']} {dane['waluta']}"), ln=True)
    
    pdf.set_xy(10, sy)
    pdf.set_font("Arial", 'B', 7.5); pdf.set_text_color(100, 100, 100); pdf.cell(55, 4.5, pdf_sanitize("CARGO TYPE / RODZAJ TOWARU:"), border=0)
    pdf.set_font("Arial", 'B', 9); pdf.set_text_color(40, 40, 40); pdf.set_xy(65, sy); pdf.multi_cell(50, 4.5, pdf_sanitize("Exhibition Structures / AV Equipment"))
    
    pdf.set_xy(10, pdf.get_y() + 1)
    pdf.set_font("Arial", 'B', 7.5); pdf.set_text_color(100, 100, 100); pdf.cell(55, 4.5, pdf_sanitize("GROSS WEIGHT / WAGA BRUTTO:"), border=0)
    pdf.set_font("Arial", 'B', 9); pdf.set_text_color(40, 40, 40); pdf.set_xy(65, pdf.get_y()); pdf.cell(50, 4.5, pdf_sanitize(f"{dane['waga']} kg"))

    pdf.set_xy(10, pdf.get_y() + 5)
    pdf.set_font("Arial", 'B', 7.5); pdf.set_text_color(100, 100, 100); pdf.cell(55, 4.5, pdf_sanitize("PAYMENT / PŁATNOŚĆ:"), border=0)
    pdf.set_font("Arial", 'B', 9); pdf.set_text_color(40, 40, 40); pdf.set_xy(65, pdf.get_y()); pdf.cell(50, 4.5, pdf_sanitize(f"{dane['termin_dni']} dni / days ({dane['data_platnosci']})"))
    
    pdf.set_xy(10, max(pdf.get_y() + 5, sy + 25))
    draw_section_header(4, "SPECIAL PROVISIONS / UWAGI SPECJALNE")
    pdf.set_font("Arial", 'I', 9)
    pdf.multi_cell(0, 4.5, pdf_sanitize(dane['uwagi']))

    return bytes(pdf.output(dest='S').encode('latin1'))

def safe_set_cell(sheet, coordinate, value):
    target_coord = coordinate
    try:
        match = re.match(r"([A-Z]+)([0-9]+)", coordinate)
        if match:
            col_str, row_str = match.groups()
            row = int(row_str)
            col = 0
            for char in col_str:
                col = col * 26 + (ord(char) - ord('A') + 1)
                
            for cr in sheet.merged_cells.ranges:
                if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
                    target_coord = str(cr).split(':')[0]
                    break
    except Exception:
        pass
    sheet[target_coord] = value

def generate_cmr_excel(dane):
    szablon_path = "Szablon_CMR.xlsx"
    if not os.path.exists(szablon_path): 
        raise FileNotFoundError(f"Brak pliku szablonu: {szablon_path}")
        
    wb = openpyxl.load_workbook(szablon_path)
    nadawca_tekst = "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182"
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=75, min_col=1, max_col=24):
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    if str(cell.fill.start_color.rgb) in ['00000000', 'FF000000', '000000']:
                        cell.fill = PatternFill(fill_type=None)

        # Standardowe pola CMR z lewej strony
        safe_set_cell(sheet, 'D6', nadawca_tekst)
        safe_set_cell(sheet, 'D14', dane.get('odbiorca', ''))
        safe_set_cell(sheet, 'D20', dane.get('miejsce_przeznaczenia', ''))
        safe_set_cell(sheet, 'D24', dane.get('data_zal', ''))
        safe_set_cell(sheet, 'H24', dane.get('miasto_zal', ''))
        safe_set_cell(sheet, 'D33', dane.get('opis_ladunku', 'MULTIMEDIA / Exhibition Equipment'))
        safe_set_cell(sheet, 'Q38', dane.get('waga', 0))
        safe_set_cell(sheet, 'E69', dane.get('miasto_zal', ''))
        safe_set_cell(sheet, 'H69', dane.get('data_zal', ''))
        safe_set_cell(sheet, 'T6', dane.get('nr_cmr', ''))
        
        # Prawa strona (Rubryka 16) - Przewoźnik, Auto, Kierowca
        # Wszystko ląduje idealnie w jednym pionowym bloku, bez dublowania na formularzu.
        safe_set_cell(sheet, 'L13', dane.get('przewoznik', ''))
        safe_set_cell(sheet, 'L14', dane.get('auto', ''))
        safe_set_cell(sheet, 'L15', dane.get('kierowca', ''))
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def get_cmr_city_format(place_name, manual_addr, df):
    if place_name == "Magazyn SQM Komorniki": 
        return "Komorniki, PL"
    if place_name == "INNE (wpisz ręcznie)":
        return ", ".join(p.strip() for p in manual_addr.split(',')[-2:]) if ',' in manual_addr else manual_addr
        
    if df is not None and not df.empty:
        row = df[df['Nazwa do listy'] == place_name]
        if not row.empty:
            r = row.iloc[0]
            miasto = str(r.get('Miasto', place_name)).strip()
            kraj = str(r.get('Kraj', '')).strip()
            skrot = str(r.get('Skrót Kraju', '')).strip() if 'Skrót Kraju' in df.columns else ''
            mapa_krajow = {
                "Polska": "PL", "Niemcy": "DE", "Francja": "FR", "Hiszpania": "ES",
                "Włochy": "IT", "Holandia": "NL", "Belgia": "BE", "Szwecja": "SE",
                "Turcja": "TR", "Szwajcaria": "CH", "Austria": "AT", "Czechy": "CZ",
                "Słowacja": "SK", "Wielka Brytania": "GB", "USA": "US"
            }
            final_kraj = skrot
            if not final_kraj or final_kraj == 'nan':
                if kraj in mapa_krajow: final_kraj = mapa_krajow[kraj]
                elif len(kraj) == 2: final_kraj = kraj.upper()
                elif kraj and kraj != 'nan': final_kraj = kraj[:2].upper()
                else: final_kraj = "PL"
            if miasto == 'nan' or not miasto: miasto = place_name
            return f"{miasto}, {final_kraj}"
    return place_name

def odtworz_dane_zlecenia(r, df_miejsca, df_przewoznicy, idx_pd, row_idx):
    nr_zlecenia = str(r.get('Numer zlecenia', r.iloc[1] if len(r) > 1 else ''))
    podpis = "".join([c for c in nr_zlecenia.split("/")[-1] if c.isalpha()])[:2] if "/" in nr_zlecenia else "PD"

    data_zal = str(r.get('Data załadunku', r.iloc[6] if len(r) > 6 else ''))
    data_roz = str(r.get('Data rozładunku', r.iloc[7] if len(r) > 7 else ''))
    data_platnosci = str(r.get('Data płatności (szacowana)', r.get('Data_Platnosci', r.iloc[9] if len(r) > 9 else '')))
    
    stawka_str = str(r.get('Stawka', r.iloc[17] if len(r) > 17 else '0 EUR'))
    stawka_final, waluta = 0.0, "EUR"
    if " " in stawka_str:
        try:
            stawka_final = float(stawka_str.split(" ")[0])
            waluta = stawka_str.split(" ")[1]
        except: pass
    else:
        try: stawka_final = float(stawka_str)
        except: pass
        
    nazwa_przewoznika = str(r.get('Zleceniobiorca', r.iloc[3] if len(r) > 3 else ''))
    detale_przewoznika = nazwa_przewoznika
    if not df_przewoznicy.empty and 'Skrócona Nazwa' in df_przewoznicy.columns:
        r_p = df_przewoznicy[df_przewoznicy['Skrócona Nazwa'] == nazwa_przewoznika]
        if not r_p.empty:
            rp_row = r_p.iloc[0]
            detale_przewoznika = f"{str(rp_row.get('Pełna Nazwa', ''))}\n{str(rp_row.get('Ulica i numer', ''))}\n{str(rp_row.get('Kod pocztowy i Miasto', ''))}, {str(rp_row.get('Kraj', 'Polska'))}\nNIP: {str(rp_row.get('NIP', ''))}".strip()
            
    z_sel = str(r.get('Miejsce Zaladunku', r.iloc[4] if len(r) > 4 else ''))
    def build_full_address(place_name, df):
        if place_name == "Magazyn SQM Komorniki":
            return "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182"
        if df is not None and not df.empty:
            row_df = df[df['Nazwa do listy'] == place_name]
            if not row_df.empty:
                r_m = row_df.iloc[0]
                return f"{r_m.get('Nazwa pełna / Firma', place_name)}\n{r_m.get('Ulica i numer', '')}\n{r_m.get('Kod pocztowy', '')} {r_m.get('Miasto', '')}, {r_m.get('Kraj', '')}"
        return place_name

    full_zal_pdf = build_full_address(z_sel, df_miejsca)
    m_roz_baza = str(r.get('Miejsce Rozladunku', r.iloc[5] if len(r) > 5 else ''))
    roz_list = m_roz_baza.split(" | ")
    lista_roz_pdf = [build_full_address(x, df_miejsca) for x in roz_list]
    if len(lista_roz_pdf) > 1: full_roz_pdf = "\n\n".join([f"DROP {idx+1}:\n{tekst}" for idx, tekst in enumerate(lista_roz_pdf)])
    else: full_roz_pdf = lista_roz_pdf[0] if lista_roz_pdf else ""
        
    uwagi_baza = str(r.get('Uwagi / Instrukcje', r.iloc[13] if len(r) > 13 else ''))
    c_auto_full, val_instrukcje, waga_val, postoj_val, wartosc_towaru = "", uwagi_baza, 1000, 0.0, 100000
    data_emp_in_1, data_emp_in_2, data_dostawa_pustych, data_odbior_pelnych, cykl_part = "", "", "", "", ""
    odbiorca_cmr_hist = full_roz_pdf
    
    if " || " in uwagi_baza:
        parts = uwagi_baza.split(" || ")
        val_instrukcje = parts[-1]
        if "%%CMR:SQM%%" in val_instrukcje:
            odbiorca_cmr_hist = "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182"
            val_instrukcje = val_instrukcje.replace(" %%CMR:SQM%%", "").replace("%%CMR:SQM%%", "")
        for p in parts:
            p = p.strip()
            if p.startswith("AUTO:"): c_auto_full = p.replace("AUTO:", "").strip()
            elif p.startswith("WART:"):
                try: wartosc_towaru = int(re.sub(r'[^0-9]', '', p))
                except: pass
            elif p.startswith("WAGA:"):
                try: waga_val = int(re.sub(r'[^0-9]', '', p))
                except: pass
            elif p.startswith("POSTOJ:"):
                try: postoj_val = float(re.sub(r'[^0-9.]', '', p))
                except: pass
            elif p.startswith("CYKL:"): cykl_part = p
    else:
        if "%%CMR:SQM%%" in uwagi_baza:
            odbiorca_cmr_hist = "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182"
            uwagi_baza = uwagi_baza.replace(" %%CMR:SQM%%", "")
        if "AUTO: " in uwagi_baza:
            try: c_auto_full = uwagi_baza.split("AUTO: ")[1].split(" ||")[0]
            except: pass

    if "EMP:" in cykl_part:
        try:
            emp_raw = cykl_part.split("EMP: ")[1].split(" | ")[0]
            if "," in emp_raw: data_emp_in_1, data_emp_in_2 = emp_raw.split(",")[0].strip(), emp_raw.split(",")[1].strip()
            else: data_emp_in_1 = emp_raw.strip()
        except: pass
    if "DEM:" in cykl_part:
        try:
            dem_raw = cykl_part.split("DEM: ")[1].split(" | ")[0]
            if "," in dem_raw: data_dostawa_pustych, data_odbior_pelnych = dem_raw.split(",")[0].strip(), dem_raw.split(",")[1].strip()
            else: data_odbior_pelnych = dem_raw.strip()
        except: pass
    elif "POWRÓT:" in cykl_part:
        try: data_odbior_pelnych = cykl_part.split("POWRÓT: ")[1].split(" | ")[0].strip()
        except: pass

    typ_zlecenia = "Pełny event" if "TARGI" in str(r.get('Typ', r.iloc[16] if len(r)>16 else '')) or "CYKL:" in uwagi_baza else "Tylko dostawa"
    uwagi_na_pdf = f"VEHICLE/DRIVER: {c_auto_full}\n{val_instrukcje}"
    
    termin_dni = 30
    try: termin_dni = (datetime.strptime(data_platnosci, "%Y-%m-%d") - datetime.strptime(data_roz.split(",")[-1].strip(), "%Y-%m-%d")).days
    except: pass
    
    paczka_pdf = {
        "typ_zlecenia": typ_zlecenia, "nr": nr_zlecenia, "przewoznik_nazwa": nazwa_przewoznika, "przewoznik_detale": detale_przewoznika,
        "stawka": stawka_final, "waluta": waluta, "postoj": postoj_val, "zaladunek": full_zal_pdf, "data_zal": data_zal,
        "rozladunek": full_roz_pdf, "data_roz": data_roz, "data_emp_in_1": data_emp_in_1, "data_emp_in_2": data_emp_in_2, 
        "data_dostawa_pustych": data_dostawa_pustych, "data_odbior_pelnych": data_odbior_pelnych, "waga": waga_val, 
        "auto": c_auto_full, "uwagi": uwagi_na_pdf, "opiekun": podpis, "termin_dni": termin_dni, "data_platnosci": data_platnosci
    }
    
    auto_val, kierowca_val = c_auto_full, ""
    if "/" in c_auto_full: auto_val, kierowca_val = c_auto_full.split("/", 1)[0].strip(), c_auto_full.split("/", 1)[1].strip()
        
    miasto_zal_val = get_cmr_city_format(z_sel, "", df_miejsca)
    
    numer_cmr_final = str(r.get('Nr_CMR', r.iloc[18] if len(r) > 18 else ''))
    if not numer_cmr_final.strip() or numer_cmr_final in ["nan", "None"]:
        numer_cmr_final = str(db.get_next_cmr_number())
        wiersz_lista = r.copy()
        if 'sheet_row' in wiersz_lista: wiersz_lista = wiersz_lista.drop('sheet_row')
        w_list = [str(x) if not pd.isna(x) else "" for x in wiersz_lista.tolist()]
        if len(w_list) < 19: w_list.extend([""] * (19 - len(w_list)))
        w_list[18] = numer_cmr_final
        db.update_row("Zlecenia", row_idx, w_list)

    dane_cmr = {
        "odbiorca": odbiorca_cmr_hist, "miejsce_przeznaczenia": full_roz_pdf, "data_zal": data_zal, "miasto_zal": miasto_zal_val,
        "opis_ladunku": "MULTIMEDIA / Exhibition Equipment", "waga": waga_val, "nr_cmr": numer_cmr_final,
        "auto": auto_val, "kierowca": kierowca_val, "przewoznik": nazwa_przewoznika
    }

    # ==========================================
    # LOGIKA CMR POWROTNEGO DLA HISTORII ZLECEŃ PRO
    # ==========================================
    last_roz = m_roz_baza.split(" | ")[-1] if m_roz_baza else ""
    miasto_zal_powrot = get_cmr_city_format(last_roz, "", df_miejsca)

    dane_cmr_powrot = {
        "odbiorca": "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182",
        "miejsce_przeznaczenia": full_zal_pdf, 
        "data_zal": str(data_odbior_pelnych) if data_odbior_pelnych else "",
        "miasto_zal": miasto_zal_powrot,
        "opis_ladunku": "MULTIMEDIA / Exhibition Equipment",
        "waga": waga_val,
        "nr_cmr": numer_cmr_final,
        "auto": auto_val,
        "kierowca": kierowca_val,
        "przewoznik": nazwa_przewoznika
    }

    return paczka_pdf, dane_cmr, nr_zlecenia, dane_cmr_powrot

@st.cache_data(ttl=30, show_spinner=False)
def pobierz_dane_z_bazy():
    return db.fetch_data("Projekty"), db.fetch_data("Miejsca"), db.fetch_data("Zleceniobiorcy"), db.fetch_data("Zlecenia")

def render(sh):
    if 'dokumenty_wygenerowane' not in st.session_state:
        st.session_state.dokumenty_wygenerowane = False
        st.session_state.pdf_bytes, st.session_state.cmr_bytes, st.session_state.cmr_powrot_bytes = None, None, None
        st.session_state.nazwa_pdf, st.session_state.nazwa_cmr, st.session_state.komunikat = "", "", ""
        
    if 'hist_gen_row' not in st.session_state:
        st.session_state.hist_gen_row = None
        st.session_state.hist_pdf_bytes, st.session_state.hist_cmr_bytes, st.session_state.hist_cmr_powrot_bytes = None, None, None
        st.session_state.hist_nr, st.session_state.hist_cmr_nr = "", ""

    st.markdown('<div class="module-header-container"><h1 class="module-title">Generator Zleceń PRO</h1><div class="module-subtitle">オーダーの生成 ✦ TRANSPORT ORDERS</div></div>', unsafe_allow_html=True)

    with st.spinner("Ładowanie telemetrii bazy danych..."):
        df_projekty, df_miejsca, df_przewoznicy, df_zlecenia = pobierz_dane_z_bazy()

    tab1, tab2 = st.tabs(["📝 Formularz / Generator PDF", "📂 Baza Zleceń PRO"])

    with tab1:
        c1, c2 = st.columns(2)
        with c1: tryb_pracy = st.radio("Wybierz tryb pracy:", ["Nowe Zlecenie", "Edycja Istniejącego Zlecenia"], horizontal=True)
        with c2: kategoria_zlecenia = st.radio("Kategoria zlecenia (gdzie zapisać?):", ["Zlecenie Poboczne (Eksport do rejestru)", "Zlecenie Eventowe (Pomiń rejestr poboczny)"], horizontal=True)
        st.markdown("<br>", unsafe_allow_html=True)

        wybrane_zlecenie_nr, gs_row_index = None, None
        val_typ_zlecenia, val_waga, val_postoj = "Tylko dostawa", 1000, 0.0
        val_data_zal, val_data_roz_1, val_data_roz_2 = datetime.now().date(), datetime.now().date(), None
        val_data_emp_in_1, val_data_emp_in_2, val_data_dostawa_pustych, val_data_odbior_pelnych = datetime.now().date(), None, datetime.now().date(), datetime.now().date()
        val_termin_dni, val_zrodlo, val_nazwa_przewoznika, val_detale_przewoznika = 30, "Przewoźnik stały (Baza)", "Wybierz...", ""
        val_stawka_final, val_waluta, val_projekt, val_z_sel = 0.0, "EUR", "Brak", "Magazyn SQM Komorniki"
        val_z_man, val_c_auto_nr, val_c_kierowca, val_wartosc_towaru = "", "", "", 100000
        val_instrukcje = "Parking strzeżony, pasy zabezpieczające; załadować po długości, casy nie mogą leżeć, kłódka / Guarded parking, safety belts; load lengthwise, cases cannot lie down, safe lock."
        val_podpis, val_miejsca_rozladunku_raw, val_odbiorca_cmr = "PD", [], "Miejsce przeznaczenia (Klient)"

        df_cargo = df_zlecenia[df_zlecenia['Dział'] == 'LOGISTYKA CARGO'].copy() if not df_zlecenia.empty and 'Dział' in df_zlecenia.columns else df_zlecenia.copy() if not df_zlecenia.empty else pd.DataFrame()

        if tryb_pracy == "Edycja Istniejącego Zlecenia":
            if not df_cargo.empty:
                wybrane_zlecenie_nr = st.selectbox("🎯 Wybierz numer zlecenia do korekty/edycji:", df_cargo['Numer zlecenia'].astype(str).tolist())
                idx_pd = df_zlecenia[df_zlecenia['Numer zlecenia'] == wybrane_zlecenie_nr].index[0]
                r_edit = df_zlecenia.iloc[idx_pd]
                gs_row_index = int(idx_pd) + 2 
                
                nr_cmr_zapisany = str(r_edit.get('Nr_CMR', r_edit.iloc[18] if len(r_edit)>18 else ''))
                if not nr_cmr_zapisany.strip() or nr_cmr_zapisany in ["nan", "None"]: nr_cmr_zapisany = db.get_next_cmr_number()
                val_typ_zlecenia = "Pełny event" if "TARGI" in str(r_edit.get('Typ', '')) or "CYKL:" in str(r_edit.get('Uwagi / Instrukcje', '')) else "Tylko dostawa"
                
                try: val_data_zal = datetime.strptime(str(r_edit.get('Data załadunku', r_edit.iloc[6])), "%Y-%m-%d").date()
                except: pass
                
                roz_str = str(r_edit.get('Data rozładunku', r_edit.iloc[7])).strip()
                if "," in roz_str:
                    parts = [p.strip() for p in roz_str.split(",")]
                    try: val_data_roz_1 = datetime.strptime(parts[0], "%Y-%m-%d").date()
                    except: pass
                    if len(parts) > 1:
                        try: val_data_roz_2 = datetime.strptime(parts[1], "%Y-%m-%d").date()
                        except: pass
                else:
                    try: val_data_roz_1 = datetime.strptime(roz_str, "%Y-%m-%d").date()
                    except: pass
                
                stawka_str = str(r_edit.get('Stawka', '0 EUR'))
                if " " in stawka_str:
                    try: val_stawka_final, val_waluta = float(stawka_str.split(" ")[0]), stawka_str.split(" ")[1]
                    except: pass
                else:
                    try: val_stawka_final = float(stawka_str)
                    except: pass
                    
                val_nazwa_przewoznika, val_projekt, val_z_sel = str(r_edit.get('Zleceniobiorca', '')), str(r_edit.get('ID Projektu', '')), str(r_edit.get('Miejsce Zaladunku', ''))
                val_miejsca_rozladunku_raw = str(r_edit.get('Miejsce Rozladunku', '')).split(" | ")
                uwagi_baza = str(r_edit.get('Uwagi / Instrukcje', r_edit.iloc[13] if len(r_edit)>13 else ''))
                
                if " || " in uwagi_baza:
                    parts = uwagi_baza.split(" || ")
                    val_instrukcje = parts[-1]
                    if "%%CMR:SQM%%" in val_instrukcje:
                        val_odbiorca_cmr = "SQM (Wysyłka na własne stoisko/event)"
                        val_instrukcje = val_instrukcje.replace(" %%CMR:SQM%%", "").replace("%%CMR:SQM%%", "")
                    for p in parts:
                        p = p.strip()
                        if p.startswith("AUTO:"):
                            auto_full = p.replace("AUTO:", "").strip()
                            if "/" in auto_full: val_c_auto_nr, val_c_kierowca = auto_full.split("/", 1)[0].strip(), auto_full.split("/", 1)[1].strip()
                            else: val_c_auto_nr = auto_full
                        elif p.startswith("WART:"):
                            try: val_wartosc_towaru = int(re.sub(r'[^0-9]', '', p))
                            except: pass
                        elif p.startswith("WAGA:"):
                            try: val_waga = int(re.sub(r'[^0-9]', '', p))
                            except: pass
                        elif p.startswith("POSTOJ:"):
                            try: val_postoj = float(re.sub(r'[^0-9.]', '', p))
                            except: pass
                        elif p.startswith("CYKL:"):
                            if "EMP:" in p:
                                try:
                                    emp_raw = p.split("EMP: ")[1].split(" | ")[0]
                                    if "," in emp_raw:
                                        if emp_raw.split(",")[0].strip(): val_data_emp_in_1 = datetime.strptime(emp_raw.split(",")[0].strip(), "%Y-%m-%d").date()
                                        if emp_raw.split(",")[1].strip(): val_data_emp_in_2 = datetime.strptime(emp_raw.split(",")[1].strip(), "%Y-%m-%d").date()
                                    elif emp_raw.strip(): val_data_emp_in_1 = datetime.strptime(emp_raw.strip(), "%Y-%m-%d").date()
                                except: pass
                            if "DEM:" in p:
                                try:
                                    dem_raw = p.split("DEM: ")[1].split(" | ")[0]
                                    if "," in dem_raw:
                                        if dem_raw.split(",")[0].strip(): val_data_dostawa_pustych = datetime.strptime(dem_raw.split(",")[0].strip(), "%Y-%m-%d").date()
                                        if dem_raw.split(",")[1].strip(): val_data_odbior_pelnych = datetime.strptime(dem_raw.split(",")[1].strip(), "%Y-%m-%d").date()
                                    elif dem_raw.strip(): val_data_odbior_pelnych = datetime.strptime(dem_raw.strip(), "%Y-%m-%d").date()
                                except: pass
                            elif "POWRÓT:" in p:
                                try:
                                    if p.split("POWRÓT: ")[1].split(" | ")[0].strip(): val_data_odbior_pelnych = datetime.strptime(p.split("POWRÓT: ")[1].split(" | ")[0].strip(), "%Y-%m-%d").date()
                                except: pass
                else:
                    if "%%CMR:SQM%%" in uwagi_baza:
                        val_odbiorca_cmr = "SQM (Wysyłka na własne stoisko/event)"
                        uwagi_baza = uwagi_baza.replace(" %%CMR:SQM%%", "")
                    if "AUTO: " in uwagi_baza:
                        try: 
                            auto_full = uwagi_baza.split("AUTO: ")[1].split(" ||")[0]
                            if "/" in auto_full: val_c_auto_nr, val_c_kierowca = auto_full.split("/", 1)[0].strip(), auto_full.split("/", 1)[1].strip()
                            else: val_c_auto_nr = auto_full.strip()
                        except: pass
                    if "WART: " in uwagi_baza:
                        try: val_wartosc_towaru = int(uwagi_baza.split("WART: ")[1].split(" PLN")[0])
                        except: pass

                try:
                    dz_roz_ost = val_data_roz_2 if val_data_roz_2 else val_data_roz_1
                    dt_plat_str = str(r_edit.get('Data płatności (szacowana)', r_edit.iloc[9] if len(r_edit)>9 else ''))
                    val_termin_dni = (datetime.strptime(dt_plat_str, "%Y-%m-%d").date() - dz_roz_ost).days
                except: val_termin_dni = 30
                    
                if "/" in wybrane_zlecenie_nr:
                    try: val_podpis = "".join([c for c in wybrane_zlecenie_nr.split("/")[-1] if c.isalpha()])[:2]
                    except: pass
                    
                if not df_przewoznicy.empty and 'Skrócona Nazwa' in df_przewoznicy.columns:
                    r_p = df_przewoznicy[df_przewoznicy['Skrócona Nazwa'] == val_nazwa_przewoznika]
                    if not r_p.empty:
                        r = r_p.iloc[0]
                        val_detale_przewoznika = f"{str(r.get('Pełna Nazwa', ''))}\n{str(r.get('Ulica i numer', ''))}\n{str(r.get('Kod pocztowy i Miasto', ''))}, {str(r.get('Kraj', 'Polska'))}\nNIP: {str(r.get('NIP', ''))}".strip()
                        val_zrodlo = "Przewoźnik stały (Baza)"
                    else:
                        val_zrodlo = "Przewoźnik z giełdy (Jednorazowy)"
                        val_detale_przewoznika = "Zweryfikuj dane adresowe z giełdy..."
            else:
                st.warning("Baza zleceń jest pusta - brak danych do edycji."); st.stop()
        else:
            nr_cmr_zapisany = db.get_next_cmr_number()

        with st.expander("🔍 Przeglądaj i wyszukaj miejsca z Bazy Lokalizacji"):
            wyszukiwana_fraza = st.text_input("Wpisz szukaną frazę (nazwa, miasto, ulica, kod):", placeholder="np. Messe Berlin...")
            if not df_miejsca.empty:
                if wyszukiwana_fraza:
                    maska = df_miejsca.astype(str).apply(lambda row: row.str.contains(wyszukiwana_fraza, case=False, na=False).any(), axis=1)
                    st.dataframe(df_miejsca[maska], use_container_width=True, hide_index=True)
                else: st.dataframe(df_miejsca, use_container_width=True, hide_index=True)

        with st.expander("➕ Brak miejsca na liście? Dodaj nową lokalizację do Słownika"):
            with st.form("form_nowe_miejsce", clear_on_submit=True):
                nowa_nazwa_lista = st.text_input("Nazwa skrócona (do listy wyboru):*", placeholder="np. BERLIN, DE - Messe Berlin")
                nowa_firma = st.text_input("Pełna nazwa / Firma:")
                nowa_ulica = st.text_input("Ulica i numer:")
                nowy_kod = st.text_input("Kod pocztowy:")
                nowe_miasto = st.text_input("Miasto:")
                k1, k2 = st.columns(2)
                nowy_kraj = k1.text_input("Kraj:", value="Polska")
                nowy_skrot = k2.text_input("Skrót Kraju (do CMR):", value="PL")
                if st.form_submit_button("💾 Zapisz lokalizację w bazie"):
                    if nowa_nazwa_lista.strip():
                        kolumny_miejsca = df_miejsca.columns.tolist() if not df_miejsca.empty else ["Nazwa do listy", "Nazwa pełna / Firma", "Ulica i numer", "Kod pocztowy", "Miasto", "Kraj", "Skrót Kraju"]
                        slownik_nowego = {"Nazwa do listy": nowa_nazwa_lista.strip(), "Nazwa pełna / Firma": nowa_firma.strip(), "Ulica i numer": nowa_ulica.strip(), "Kod pocztowy": nowy_kod.strip(), "Miasto": nowe_miasto.strip(), "Kraj": nowy_kraj.strip(), "Skrót Kraju": nowy_skrot.strip()}
                        nowy_wiersz = [str(slownik_nowego.get(kol, "")) for kol in kolumny_miejsca]
                        if db.append_data("Miejsca", nowy_wiersz):
                            st.success(f"✅ Dodano pomyślnie: {nowa_nazwa_lista}")
                            st.cache_data.clear(); st.rerun()

        lista_eventow = df_projekty['Nazwa Eventu'].dropna().unique().tolist() if not df_projekty.empty else ["Brak"]
        lista_miejsc_baza = df_miejsca['Nazwa do listy'].tolist() if not df_miejsca.empty else []
        opcje_lokalizacji = ["Magazyn SQM Komorniki"] + lista_miejsc_baza + ["INNE (wpisz ręcznie)"]

        typ_zlecenia = st.radio("Tryb operacji:", ["Tylko dostawa", "Pełny event"], index=["Tylko dostawa", "Pełny event"].index(val_typ_zlecenia), horizontal=True)

        with st.container(border=True):
            st.markdown("<p style='color: #C5A880; font-weight: 700; margin-bottom: 5px;'>1. Harmonogram Zlecenia</p>", unsafe_allow_html=True)
            waga = st.number_input("Waga ładunku (kg):", min_value=100, step=100, value=int(val_waga))
            d1, d2, d3 = st.columns(3)
            data_zal = d1.date_input("Data załadunku (PL):", val_data_zal)
            data_roz_1 = d2.date_input("Rozładunek 1 (Cel):", val_data_roz_1)
            data_roz_2 = d3.date_input("Rozładunek 2 (Opcja):", value=val_data_roz_2)
            data_roz_combined = str(data_roz_1)
            if data_roz_2: data_roz_combined += f", {data_roz_2}"
            
            if typ_zlecenia == "Pełny event":
                st.markdown("<hr style='margin: 10px 0; border-color: rgba(197, 168, 128, 0.1);'>", unsafe_allow_html=True)
                st.markdown("<p style='font-size: 13px; color: #8C8477; margin-bottom: 5px;'>📦 Odbiór pustych skrzyń po rozładunku (Empties In):</p>", unsafe_allow_html=True)
                e1, e2 = st.columns(2)
                data_emp_in_1 = e1.date_input("Data odbioru 1:", val_data_emp_in_1)
                data_emp_in_2 = e2.date_input("Data odbioru 2 (Opcjonalnie):", value=val_data_emp_in_2)
                st.markdown("<p style='font-size: 13px; color: #8C8477; margin-top: 10px; margin-bottom: 5px;'>🛠️ Demontaż targów (Powrót):</p>", unsafe_allow_html=True)
                r1, r2 = st.columns(2)
                data_dostawa_pustych = r1.date_input("Dostawa pustych casów:", val_data_dostawa_pustych)
                data_odbior_pelnych = r2.date_input("Odbiór pełnych po demontażu:", val_data_odbior_pelnych)
            else: data_emp_in_1, data_emp_in_2, data_dostawa_pustych, data_odbior_pelnych = "", "", "", ""

        with st.container(border=True):
            st.markdown("<p style='color: #C5A880; font-weight: 700; margin-bottom: 5px;'>2. Wybór Przewoźnika i Płatności</p>", unsafe_allow_html=True)
            zrodlo = st.radio("Sposób wyboru podwykonawcy:", ["Przewoźnik stały (Baza)", "Przewoźnik z giełdy (Jednorazowy)"], index=["Przewoźnik stały (Baza)", "Przewoźnik z giełdy (Jednorazowy)"].index(val_zrodlo) if val_zrodlo in ["Przewoźnik stały (Baza)", "Przewoźnik z giełdy (Jednorazowy)"] else 0, horizontal=True)
            detale_przewoznika, nazwa_przewoznika = "", ""

            if zrodlo == "Przewoźnik stały (Baza)":
                f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
                lista_cennikowa = df_przewoznicy['Skrócona Nazwa'].dropna().tolist() if not df_przewoznicy.empty else []
                if tryb_pracy == "Edycja Istniejącego Zlecenia" and val_nazwa_przewoznika not in lista_cennikowa: lista_cennikowa.append(val_nazwa_przewoznika)
                nazwa_przewoznika = f1.selectbox("Wybierz partnera ze słownika:", ["Wybierz..."] + lista_cennikowa, index=lista_cennikowa.index(val_nazwa_przewoznika)+1 if val_nazwa_przewoznika in lista_cennikowa else 0)
                if nazwa_przewoznika != "Wybierz...":
                    if not df_przewoznicy.empty and 'Skrócona Nazwa' in df_przewoznicy.columns:
                        row_p = df_przewoznicy[df_przewoznicy['Skrócona Nazwa'] == nazwa_przewoznika]
                        if not row_p.empty:
                            r = row_p.iloc[0]
                            detale_przewoznika = f"{str(r.get('Pełna Nazwa', nazwa_przewoznika))}\n{str(r.get('Ulica i numer', ''))}\n{str(r.get('Kod pocztowy i Miasto', ''))}, {str(r.get('Kraj', 'Polska'))}\nNIP: {str(r.get('NIP', ''))}".strip()
                        else: detale_przewoznika = nazwa_przewoznika
                    else: detale_przewoznika = nazwa_przewoznika

                stawka_final = f2.number_input("Stawka Total:", value=float(val_stawka_final))
                waluta = f3.selectbox("Waluta:", ["EUR", "PLN"], index=["EUR", "PLN"].index(val_waluta) if val_waluta in ["EUR", "PLN"] else 0)
                postoj = f4.number_input("Postój:", value=float(val_postoj)) if typ_zlecenia == "Pełny event" else 0.0
            else:
                nazwa_przewoznika = st.text_input("Nazwa firmy z giełdy:", value=val_nazwa_przewoznika)
                detale_przewoznika = st.text_area("Pełne dane (Adres, NIP do zlecenia):", value=val_detale_przewoznika)
                f1, f2, f3 = st.columns(3)
                stawka_final = f1.number_input("Stawka netto:", min_value=0.0, value=float(val_stawka_final))
                waluta = f2.selectbox("Waluta:", ["EUR", "PLN"], index=["EUR", "PLN"].index(val_waluta) if val_waluta in ["EUR", "PLN"] else 0)
                postoj = f3.number_input("Postój:", min_value=0.0, value=float(val_postoj)) if typ_zlecenia == "Pełny event" else 0.0
                
            t1, t2 = st.columns([1, 2])
            termin_dni = t1.number_input("Termin płatności (dni):", min_value=0, max_value=120, value=int(val_termin_dni), step=1)
            
            if typ_zlecenia == "Pełny event" and data_odbior_pelnych:
                data_platnosci = data_odbior_pelnych + timedelta(days=2) + timedelta(days=termin_dni)
                t2.info(f"📅 Wyliczona data zapłaty (Odbiór z targów + 2 dni drogi + {termin_dni} dni): **{data_platnosci.strftime('%d.%m.%Y')}**")
            else:
                data_platnosci = (data_roz_2 if data_roz_2 else data_roz_1) + timedelta(days=termin_dni)
                t2.info(f"📅 Wyliczona data zapłaty (Rozładunek + {termin_dni} dni): **{data_platnosci.strftime('%d.%m.%Y')}**")

        with st.container(border=True):
            st.markdown("<p style='color: #C5A880; font-weight: 700; margin-bottom: 5px;'>3. Logistyka Miejsc</p>", unsafe_allow_html=True)
            projekt = st.selectbox("Przypisz do Projektu (Opcjonalnie):", lista_eventow, index=lista_eventow.index(val_projekt) if val_projekt in lista_eventow else 0)
            
            l1, l2 = st.columns(2)
            with l1:
                idx_z = opcje_lokalizacji.index(val_z_sel) if val_z_sel in opcje_lokalizacji else (opcje_lokalizacji.index("INNE (wpisz ręcznie)") if "INNE (wpisz ręcznie)" in opcje_lokalizacji else 0)
                z_sel = st.selectbox("Miejsce startu (Załadunek):", opcje_lokalizacji, index=idx_z)
                z_man = st.text_input("Adres startu (ręcznie):", value=val_z_sel if val_z_sel not in opcje_lokalizacji else "") if z_sel == "INNE (wpisz ręcznie)" else ""
                
            miejsca_rozladunku = []
            with l2:
                if typ_zlecenia == "Tylko dostawa":
                    st.markdown("🚚 **Dostawa wieloetapowa (Drop)**")
                    liczba_punktow = st.number_input("Liczba miejsc rozładunku:", min_value=1, max_value=10, value=int(len(val_miejsca_rozladunku_raw) if len(val_miejsca_rozladunku_raw) > 0 else 1), step=1)
                    for i in range(int(liczba_punktow)):
                        def_r_item = val_miejsca_rozladunku_raw[i] if i < len(val_miejsca_rozladunku_raw) else "Wybierz..."
                        idx_r = opcje_lokalizacji.index(def_r_item) if def_r_item in opcje_lokalizacji else (opcje_lokalizacji.index("INNE (wpisz ręcznie)") if "INNE (wpisz ręcznie)" in opcje_lokalizacji else 0)
                        r_s = st.selectbox(f"Cel dostawy DROP {i+1}:", opcje_lokalizacji, index=idx_r, key=f"r_sel_{i}")
                        r_m = st.text_input(f"Adres DROP {i+1} (ręcznie):", value=def_r_item if def_r_item not in opcje_lokalizacji and def_r_item != "Wybierz..." else "", key=f"r_man_{i}") if r_s == "INNE (wpisz ręcznie)" else ""
                        miejsca_rozladunku.append((r_s, r_m))
                else:
                    def_r_item = val_miejsca_rozladunku_raw[0] if len(val_miejsca_rozladunku_raw) > 0 else "Wybierz..."
                    idx_r = opcje_lokalizacji.index(def_r_item) if def_r_item in opcje_lokalizacji else (opcje_lokalizacji.index("INNE (wpisz ręcznie)") if "INNE (wpisz ręcznie)" in opcje_lokalizacji else 0)
                    r_s = st.selectbox("Miejsce celu (Targi):", opcje_lokalizacji, index=idx_r)
                    r_m = st.text_input("Adres celu (ręcznie):", value=def_r_item if def_r_item not in opcje_lokalizacji and def_r_item != "Wybierz..." else "") if r_s == "INNE (wpisz ręcznie)" else ""
                    miejsca_rozladunku.append((r_s, r_m))
                    
            st.markdown("<hr style='margin: 10px 0; border-color: rgba(197, 168, 128, 0.2);'>", unsafe_allow_html=True)
            odbiorca_cmr_ui = st.radio("Kto jest formalnym Odbiorcą na dokumencie CMR (Box 2)?:", ["Miejsce przeznaczenia (Klient)", "SQM (Wysyłka na własne stoisko/event)"], index=1 if val_odbiorca_cmr == "SQM (Wysyłka na własne stoisko/event)" else 0, horizontal=True)

        with st.container(border=True):
            st.markdown("<p style='color: #C5A880; font-weight: 700; margin-bottom: 5px;'>4. Realizacja i Dodatkowe Uwagi</p>", unsafe_allow_html=True)
            col_auto, col_kier, col_wart = st.columns([1.5, 1.5, 1])
            c_auto_nr = col_auto.text_input("Nr rejestracyjny (Auto):", value=val_c_auto_nr, placeholder="np. PO 12345")
            c_kierowca = col_kier.text_input("Kierowca (Imię i Nazwisko):", value=val_c_kierowca, placeholder="np. Jan Kowalski")
            wartosc_towaru = col_wart.number_input("Wymagana Gwarancja OCP (PLN):", min_value=0, value=val_wartosc_towaru)
            u1, u2 = st.columns([3, 1])
            instrukcje = u1.text_area("Instrukcje dodatkowe na Zlecenie:", value=val_instrukcje, height=80)
            podpis = u2.radio("Podpis Koordynatora:", ["PD", "PK"], index=["PD", "PK"].index(val_podpis) if val_podpis in ["PD", "PK"] else 0, horizontal=True)

        btn_label = "⚡ ZAPISZ ZMIANY I REGENERUJ DOKUMENTY" if tryb_pracy == "Edycja Istniejącego Zlecenia" else "⚡ GENERUJ I ZAPISZ ZLECENIE PRO"

        if st.button(btn_label, type="primary", use_container_width=True):
            if not nazwa_przewoznika or nazwa_przewoznika == "Wybierz...": st.error("Wybierz lub wpisz firmę przewozową!")
            else:
                with st.spinner("Generowanie dokumentów i aktualizacja chmury..."):
                    final_zal_db = z_man if z_sel == "INNE (wpisz ręcznie)" else z_sel
                    
                    def build_full_address(place_name, manual_addr, df):
                        if place_name == "INNE (wpisz ręcznie)": return manual_addr
                        if place_name == "Magazyn SQM Komorniki": return "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182"
                        if df is not None and not df.empty:
                            row = df[df['Nazwa do listy'] == place_name]
                            if not row.empty:
                                r = row.iloc[0]
                                return f"{r.get('Nazwa pełna / Firma', place_name)}\n{r.get('Ulica i numer', '')}\n{r.get('Kod pocztowy', '')} {r.get('Miasto', '')}, {r.get('Kraj', '')}"
                        return place_name

                    full_zal_pdf = build_full_address(z_sel, z_man, df_miejsca)
                    lista_roz_db, lista_roz_pdf = [], []
                    for r_s, r_m in miejsca_rozladunku:
                        lista_roz_db.append(r_m if r_s == "INNE (wpisz ręcznie)" else r_s)
                        lista_roz_pdf.append(build_full_address(r_s, r_m, df_miejsca))
                        
                    final_roz_db = " | ".join(lista_roz_db)
                    if len(lista_roz_pdf) > 1: full_roz_pdf = "\n\n".join([f"DROP {idx+1}:\n{tekst}" for idx, tekst in enumerate(lista_roz_pdf)])
                    else: full_roz_pdf = lista_roz_pdf[0]
                    
                    c_auto_combined = f"{c_auto_nr} / {c_kierowca}" if c_auto_nr and c_kierowca else f"{c_auto_nr}{c_kierowca}"
                    
                    historia_cyklu = f"CYKL: {data_zal} -> {data_roz_combined}"
                    if typ_zlecenia == "Pełny event":
                        emp_str = str(data_emp_in_1)
                        if data_emp_in_2: emp_str += f",{data_emp_in_2}"
                        dem_str = f"{data_dostawa_pustych},{data_odbior_pelnych}"
                        historia_cyklu += f" | EMP: {emp_str} | DEM: {dem_str}"
                    
                    pelne_uwagi_db = f"AUTO: {c_auto_combined} || WART: {wartosc_towaru} PLN || WAGA: {waga} || POSTOJ: {postoj} || {historia_cyklu} || {instrukcje}"
                    if odbiorca_cmr_ui == "SQM (Wysyłka na własne stoisko/event)": pelne_uwagi_db += " %%CMR:SQM%%"
                        
                    uwagi_na_pdf = f"VEHICLE/DRIVER: {c_auto_combined}\n{instrukcje}"
                    
                    if tryb_pracy == "Edycja Istniejącego Zlecenia": nr_zlecenia = wybrane_zlecenie_nr
                    else:
                        idx = db.get_next_daily_number(datetime.now().strftime("%Y-%m-%d"))
                        prefix = "ZLP" if kategoria_zlecenia == "Zlecenie Poboczne (Eksport do rejestru)" else "EVT"
                        nr_zlecenia = f"{prefix}{datetime.now().strftime('%y/%m%d')}/{podpis}{idx:02d}"
                    
                    paczka_pdf = {
                        "typ_zlecenia": typ_zlecenia, "nr": nr_zlecenia, 
                        "przewoznik_nazwa": nazwa_przewoznika, "przewoznik_detale": detale_przewoznika,
                        "stawka": stawka_final, "waluta": waluta, "postoj": postoj,
                        "zaladunek": full_zal_pdf, "data_zal": str(data_zal),
                        "rozladunek": full_roz_pdf, "data_roz": data_roz_combined,
                        "data_emp_in_1": str(data_emp_in_1), "data_emp_in_2": str(data_emp_in_2) if data_emp_in_2 else "",
                        "data_dostawa_pustych": str(data_dostawa_pustych), "data_odbior_pelnych": str(data_odbior_pelnych),
                        "waga": waga, "auto": c_auto_combined, "uwagi": uwagi_na_pdf, "opiekun": podpis,
                        "termin_dni": termin_dni, "data_platnosci": data_platnosci.strftime('%d.%m.%Y')
                    }
                    
                    wiersz_db = [
                        str(datetime.now().strftime("%Y-%m-%d %H:%M")), str(nr_zlecenia), "LOGISTYKA CARGO", str(nazwa_przewoznika),
                        str(final_zal_db), str(final_roz_db), str(data_zal), str(data_roz_combined), "Zabudowa Targowa PRO",
                        str(data_platnosci.strftime('%d.%m.%Y')), "", "", "", str(pelne_uwagi_db), "", str(projekt), "TARGI", f"{stawka_final} {waluta}",
                        str(nr_cmr_zapisany)
                    ]
                    
                    if tryb_pracy == "Edycja Istniejącego Zlecenia":
                        operacja_sukces = db.update_row("Zlecenia", gs_row_index, wiersz_db)
                    else:
                        operacja_sukces = db.append_data("Zlecenia", wiersz_db)
                        if operacja_sukces and kategoria_zlecenia == "Zlecenie Poboczne (Eksport do rejestru)":
                            wiersz_poboczne = [
                                str(nr_zlecenia), str(nazwa_przewoznika), f"PROJEKT: {projekt} | {instrukcje}",  
                                str(data_zal), str(data_roz_combined), str(termin_dni),                       
                                str(data_platnosci.strftime('%d.%m.%Y')), "PLANOWANIE", "NIE", "NIE", "NIE"                                  
                            ]
                            db.append_data("Zlecenia Poboczne", wiersz_poboczne)
                            
                    if operacja_sukces:
                        pdf_bytes = generate_pro_pdf(paczka_pdf)
                        miasto_zal_val = get_cmr_city_format(z_sel, z_man, df_miejsca)
                        odbiorca_cmr_text = "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182" if odbiorca_cmr_ui == "SQM (Wysyłka na własne stoisko/event)" else full_roz_pdf
                        
                        dane_cmr = {
                            "odbiorca": odbiorca_cmr_text, "miejsce_przeznaczenia": full_roz_pdf, "data_zal": str(data_zal),
                            "miasto_zal": miasto_zal_val, "opis_ladunku": "MULTIMEDIA / Exhibition Equipment",
                            "waga": waga, "nr_cmr": nr_cmr_zapisany, "auto": c_auto_nr, "kierowca": c_kierowca, "przewoznik": nazwa_przewoznika
                        }
                        cmr_bytes = generate_cmr_excel(dane_cmr)
                        
                        if miejsca_rozladunku:
                            last_r_s, last_r_m = miejsca_rozladunku[-1]
                            miasto_zal_powrot = get_cmr_city_format(last_r_s, last_r_m, df_miejsca)
                        else:
                            miasto_zal_powrot = ""
                            
                        dane_cmr_powrot = {
                            "odbiorca": "SQM Prosta Spółka Akcyjna ;\nul. Poznańska 165, 62-052 Komorniki,\nNIP: 7792361182",
                            "miejsce_przeznaczenia": full_zal_pdf, 
                            "data_zal": str(data_odbior_pelnych) if data_odbior_pelnych else "",
                            "miasto_zal": miasto_zal_powrot,
                            "opis_ladunku": "MULTIMEDIA / Exhibition Equipment",
                            "waga": waga,
                            "nr_cmr": nr_cmr_zapisany,
                            "auto": c_auto_nr,
                            "kierowca": c_kierowca,
                            "przewoznik": nazwa_przewoznika
                        }
                        st.session_state.cmr_powrot_bytes = generate_cmr_excel(dane_cmr_powrot)
                        
                        st.session_state.komunikat = f"🎉 Zlecenie {nr_zlecenia} zmodyfikowane!" if tryb_pracy == "Edycja Istniejącego Zlecenia" else f"✅ Zlecenie {nr_zlecenia} wygenerowane!"
                        st.session_state.pdf_bytes, st.session_state.cmr_bytes = pdf_bytes, cmr_bytes
                        st.session_state.nazwa_pdf = f"Order_{nr_zlecenia.replace('/', '_')}.pdf"
                        st.session_state.nazwa_cmr = f"CMR_{nr_zlecenia.replace('/', '_')}_{nr_cmr_zapisany}.xlsx"
                        st.session_state.dokumenty_wygenerowane = True
                        st.cache_data.clear(); st.rerun() 

        if st.session_state.dokumenty_wygenerowane:
            st.success(st.session_state.komunikat)
            col_pdf, col_cmr, col_cmr_pow = st.columns(3)
            with col_pdf: st.download_button("📥 POBIERZ ZLECENIE (PDF)", data=st.session_state.pdf_bytes, file_name=st.session_state.nazwa_pdf, mime="application/pdf", use_container_width=True)
            with col_cmr: st.download_button(f"📝 POBIERZ CMR", data=st.session_state.cmr_bytes, file_name=st.session_state.nazwa_cmr, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with col_cmr_pow:
                if st.session_state.cmr_powrot_bytes:
                    st.download_button(f"🔙 POBIERZ CMR (POWRÓT)", data=st.session_state.cmr_powrot_bytes, file_name=f"CMR_POWROT_{st.session_state.nazwa_cmr}", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Wyczyść i przygotuj nowe zlecenie", use_container_width=True): st.session_state.dokumenty_wygenerowane = False; st.rerun()

    with tab2:
        st.markdown('<h3 style="color: #E2DCD3; font-family: \'Shippori Mincho\', serif;">Aktywne Zlecenia PRO</h3>', unsafe_allow_html=True)
        try:
            df_pro = df_zlecenia.copy()
            if not df_pro.empty:
                df_pro['sheet_row'] = df_pro.index + 2
                if 'Dział' in df_pro.columns: df_pro = df_pro[df_pro['Dział'] == 'LOGISTYKA CARGO']
                df_pro = df_pro.iloc[::-1]
                
                if df_pro.empty: st.info("Brak aktywnych zleceń PRO w bazie danych.")
                else:
                    for index, row in df_pro.iterrows():
                        nr, projekt, data_zal, miejsce_zal, miejsce_roz, przewoznik = str(row.get("Numer zlecenia", "Brak numeru")), str(row.get("ID Projektu", "---")), str(row.get("Data załadunku", "---")), str(row.get("Miejsce Zaladunku", "---")), str(row.get("Miejsce Rozladunku", "---")), str(row.get("Zleceniobiorca", "---"))
                        row_idx, idx_pd = int(row['sheet_row']), int(row.name)
                        
                        st.markdown(f"""<div class="custom-row" style="margin-bottom: 5px;"><div class="cr-col" style="flex: 2.5;"><div class="cr-title">🚚 {nr}</div><div class="cr-text" style="color: #C5A880;">📦 Projekt: <strong>{projekt}</strong></div><div class="cr-text">👤 Przewoźnik: <strong>{przewoznik}</strong></div></div><div class="cr-col" style="flex: 2;"><div class="cr-text">📅 Załadunek: {data_zal}</div><div class="cr-text">📍 Skąd: {miejsce_zal}</div><div class="cr-text">🏁 Dokąd: {miejsce_roz}</div></div></div>""", unsafe_allow_html=True)
                        
                        c_info, c_docs, c_del = st.columns([3, 1.5, 1])
                        with c_docs:
                            if st.button("📄 Przygotuj Dokumenty", key=f"doc_pro_{row_idx}", use_container_width=True):
                                with st.spinner("Rekonstrukcja danych z bazy..."):
                                    paczka, cmr, hist_nr, cmr_powrot = odtworz_dane_zlecenia(row, df_miejsca, df_przewoznicy, idx_pd, row_idx)
                                    st.session_state.hist_pdf_bytes = generate_pro_pdf(paczka)
                                    st.session_state.hist_cmr_bytes = generate_cmr_excel(cmr)
                                    st.session_state.hist_cmr_powrot_bytes = generate_cmr_excel(cmr_powrot)
                                    st.session_state.hist_nr = hist_nr
                                    st.session_state.hist_cmr_nr = cmr.get("nr_cmr", "")
                                    st.session_state.hist_gen_row = row_idx
                                    st.rerun()
                        with c_del:
                            if st.button("🗑️ Usuń", key=f"del_pro_{row_idx}", use_container_width=True):
                                db.delete_row("Zlecenia", row_idx); st.success(f"Zlecenie usunięte!"); st.rerun()
                        
                        if st.session_state.hist_gen_row == row_idx:
                            st.success(f"Pliki gotowe do pobrania!")
                            d1, d2, d3 = st.columns(3)
                            with d1: st.download_button("📥 POBIERZ PDF", data=st.session_state.hist_pdf_bytes, file_name=f"Order_{st.session_state.hist_nr.replace('/','_')}.pdf", mime="application/pdf", key=f"dl_pdf_{row_idx}", use_container_width=True)
                            with d2:
                                nazwa_pliku_cmr = f"CMR_{st.session_state.hist_nr.replace('/','_')}_{st.session_state.hist_cmr_nr}.xlsx" if st.session_state.hist_cmr_nr else f"CMR_{st.session_state.hist_nr.replace('/','_')}.xlsx"
                                st.download_button("📝 POBIERZ CMR", data=st.session_state.hist_cmr_bytes, file_name=nazwa_pliku_cmr, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_cmr_{row_idx}", use_container_width=True)
                            with d3:
                                nazwa_pliku_powrot = f"CMR_POWROT_{st.session_state.hist_nr.replace('/','_')}_{st.session_state.hist_cmr_nr}.xlsx" if st.session_state.hist_cmr_nr else f"CMR_POWROT_{st.session_state.hist_nr.replace('/','_')}.xlsx"
                                st.download_button("🔙 POBIERZ CMR POWRÓT", data=st.session_state.hist_cmr_powrot_bytes, file_name=nazwa_pliku_powrot, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_cmr_powrot_{row_idx}", use_container_width=True)
                                
                        st.markdown('<hr style="border-color: rgba(197, 168, 128, 0.1); margin: 5px 0 15px 0;">', unsafe_allow_html=True)
            else: st.info("Baza PRO jest pusta.")
        except Exception as e: st.error(f"Błąd komunikacji z bazą Zleceń PRO: {e}")
